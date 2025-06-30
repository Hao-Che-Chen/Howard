import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import os
import json
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time

# 設定檔路徑
settings_file = os.path.join(os.path.expanduser("~"), "inventory_settings.json")

# 讀取最後使用的Excel路徑
def load_last_path():
    if os.path.exists(settings_file):
        with open(settings_file, "r") as f:
            return json.load(f).get("last_excel_path", "")
    return ""

def save_last_path(path):
    with open(settings_file, "w") as f:
        json.dump({"last_excel_path": path}, f)

# 初始化 Excel 格式
def init_excel(path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "入庫"
    ws1.append(["料號", "入庫數量", "日期", "時間", "入庫位置(備註)"])
    ws2 = wb.create_sheet("使用者入庫")
    # 增加入庫位置欄位
    ws2.append(["使用人", "料號", "剩餘數量", "日期", "時間", "使用數量", "入庫位置(備註)"])
    ws3 = wb.create_sheet("目前庫存")
    ws3.append(["料號", "庫存數量", "借出狀態", "借出人", "庫存位置(備註)"])
    wb.save(path)

# 匯入Excel
def import_excel():
    global excel_file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        excel_file = file_path
        save_last_path(excel_file)
        refresh_inventory()
        messagebox.showinfo("成功", "已匯入 Excel 檔案！")

# 匯出Excel
def export_excel():
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        try:
            shutil.copy(excel_file, file_path)
            messagebox.showinfo("成功", "已匯出 Excel 檔案！")
        except Exception as e:
            messagebox.showerror("匯出失敗", f"匯出時發生錯誤: {str(e)}")

# 新品入庫 - 重複使用剩餘數量為0流水號，沒有則新增流水號
def save_instock():
    part_no = entry_in_part.get().strip()
    qty_per_group = entry_in_unit_qty.get().strip()
    groups = entry_in_groups.get().strip()
    location = entry_in_location.get().strip()  # 入庫位置，可空白
    
    # 只檢查必填欄位
    if not part_no or not qty_per_group or not groups:
        messagebox.showwarning("提醒", "請填寫料號、每組數量和組數")
        return
    try:
        qty_per_group_int = int(qty_per_group)
        groups_int = int(groups)
        if qty_per_group_int <= 0 or groups_int <= 0:
            messagebox.showerror("錯誤", "數量和組數必須大於0")
            return
    except:
        messagebox.showerror("錯誤", "數量與組數必須為整數")
        return

    total_qty = qty_per_group_int * groups_int
    now = datetime.now()

    try:
        wb = load_workbook(excel_file)
        ws_in = wb["入庫"]
        ws_inv = wb["目前庫存"]

        reusable_nums = []
        existing_nums = set()

        for row in ws_inv.iter_rows(min_row=2):
            val = row[0].value
            if val and val.startswith(part_no + "-"):
                suffix = val.split("-")[-1]
                if suffix.isdigit():
                    num = int(suffix)
                    existing_nums.add(num)
                    qty = row[1].value or 0
                    if qty == 0:
                        reusable_nums.append((num, row))

        reusable_nums.sort(key=lambda x: x[0])

        used_count = 0

        # 先用掉可重用流水號
        for num, row in reusable_nums:
            if used_count >= groups_int:
                break
            full_part_no = f"{part_no}-{num}"
            ws_in.append([full_part_no, qty_per_group_int, now.strftime("%Y-%m-%d"), 
                          now.strftime("%H:%M:%S"), location])
            # 更新庫存數量，狀態清空，並設定庫存位置
            row[1].value = qty_per_group_int
            row[2].value = ""
            row[3].value = ""
            row[4].value = location  # 設定庫存位置（可空白）
            used_count += 1

        # 不足部分新增流水號
        next_num = 1
        while next_num in existing_nums:
            next_num += 1
        for i in range(groups_int - used_count):
            full_part_no = f"{part_no}-{next_num + i}"
            ws_in.append([full_part_no, qty_per_group_int, now.strftime("%Y-%m-%d"), 
                          now.strftime("%H:%M:%S"), location])
            # 新增庫存記錄，包括庫存位置（可空白）
            ws_inv.append([full_part_no, qty_per_group_int, "", "", location])

        wb.save(excel_file)
    except Exception as e:
        messagebox.showerror("錯誤", f"請關閉EXCEL檔案,存檔錯誤:\n{e}")
        return

    refresh_inventory()
    entry_in_part.delete(0, tk.END)
    entry_in_unit_qty.delete(0, tk.END)
    entry_in_groups.delete(0, tk.END)
    entry_in_location.delete(0, tk.END)
    messagebox.showinfo("完成", f"共入庫 {groups_int} 組，總數量 {total_qty} 件")

# 借出
def lend_part():
    lender = entry_lend_user.get().strip()
    part_no = entry_lend_part.get().strip()
    if not lender or not part_no:
        messagebox.showwarning("提醒", "請填寫借出人和料號")
        return

    try:
        wb = load_workbook(excel_file)
        ws_inv = wb["目前庫存"]
        found = False
        for row in ws_inv.iter_rows(min_row=2):
            if row[0].value == part_no:
                found = True
                current_qty = row[1].value or 0
                if current_qty <= 0:
                    messagebox.showerror("錯誤", f"{part_no} 庫存不足")
                    return
                if row[2].value == "借出中":
                    messagebox.showerror("錯誤", f"{part_no} 目前已借出")
                    return
                
                # 借出操作
                row[2].value = "借出中"
                row[3].value = lender
                
                wb.save(excel_file)
                refresh_inventory()
                entry_lend_user.delete(0, tk.END)
                entry_lend_part.delete(0, tk.END)
                messagebox.showinfo("完成", f"{part_no} 借出成功")
                return
        
        if not found:
            messagebox.showerror("錯誤", f"料號 {part_no} 不存在")
    except Exception as e:
        messagebox.showerror("錯誤", f"錯誤:\n{e}")

# 使用人入庫（歸還）
def save_outstock():
    user = entry_user.get().strip()
    part_no = entry_part.get().strip()
    remain_qty = entry_qty.get().strip()
    location = entry_location.get().strip()  # 入庫位置，可空白
    
    # 只檢查必填欄位
    if not user or not part_no or not remain_qty:
        messagebox.showwarning("提醒", "請填寫使用人、料號和剩餘數量")
        return
    try:
        remain = int(remain_qty)
        if remain < 0:
            messagebox.showerror("錯誤", "剩餘數量不能為負數")
            return
    except:
        messagebox.showerror("錯誤", "剩餘數量必須為整數")
        return

    now = datetime.now()
    try:
        wb = load_workbook(excel_file)
        ws_out = wb["使用者入庫"]
        ws_inv = wb["目前庫存"]

        found = False
        for row in ws_inv.iter_rows(min_row=2):
            if row[0].value == part_no:
                found = True
                # 直接更新庫存數量為輸入的剩餘數量（可以是0）
                last_qty = row[1].value or 0
                used_qty = last_qty - remain
                
                if used_qty < 0:
                    messagebox.showerror("錯誤", "剩餘數量不可大於目前庫存")
                    return

                # 記錄使用歷史（新增入庫位置，可空白）
                ws_out.append([user, part_no, remain, now.strftime("%Y-%m-%d"), 
                               now.strftime("%H:%M:%S"), used_qty, location])
                
                # 更新庫存數量
                row[1].value = remain
                
                # 更新庫存位置（如果提供了新位置）
                if location:  # 只有當位置非空白時才更新
                    row[4].value = location
                
                # 清除借出狀態（如果借出人是當前使用者）
                if row[2].value == "借出中" and row[3].value == user:
                    row[2].value = ""
                    row[3].value = ""
                break

        if not found:
            messagebox.showerror("錯誤", f"料號 {part_no} 不存在")
            return
            
        wb.save(excel_file)
    except Exception as e:
        messagebox.showerror("錯誤", f"錯誤:\n{e}")
        return

    refresh_inventory()
    entry_user.delete(0, tk.END)
    entry_part.delete(0, tk.END)
    entry_qty.delete(0, tk.END)
    entry_location.delete(0, tk.END)  # 清空入庫位置欄位
    messagebox.showinfo("完成", f"使用人入庫成功！使用數量：{used_qty}")

# 刷新目前庫存
def refresh_inventory():
    for item in tree.get_children():
        tree.delete(item)
    try:
        wb = load_workbook(excel_file)
        ws_inv = wb["目前庫存"]
        for row in ws_inv.iter_rows(min_row=2, values_only=True):
            # 確保借出狀態和借出人顯示正確
            status = row[2] if row[2] else ""
            lender = row[3] if row[3] else ""
            location = row[4] if row[4] else ""
            tree.insert("", tk.END, values=(row[0], row[1], status, lender, location))
    except Exception as e:
        messagebox.showerror("錯誤", f"讀取庫存錯誤:\n{e}")

# 查詢使用者歷史
def search_history():
    search_win = tk.Toplevel(root)
    search_win.title("查詢歷史資料")
    search_win.geometry("1100x550")  # 加大視窗以顯示新欄位

    frame = tk.Frame(search_win)
    frame.pack(pady=10)

    tk.Label(frame, text="料號:").grid(row=0, column=0, padx=5)
    search_part = tk.Entry(frame, width=20)
    search_part.grid(row=0, column=1, padx=5)

    tk.Label(frame, text="使用人:").grid(row=0, column=2, padx=5)
    search_user = tk.Entry(frame, width=20)
    search_user.grid(row=0, column=3, padx=5)

    # 入庫位置欄位
    tree_history = ttk.Treeview(search_win, columns=("使用人", "料號", "使用數量", "日期", "時間", "入庫位置"), show="headings")
    tree_history.column("使用人", width=150)
    tree_history.column("料號", width=200)
    tree_history.column("使用數量", width=100)
    tree_history.column("日期", width=150)
    tree_history.column("時間", width=150)
    tree_history.column("入庫位置", width=250)
    
    for col in ("使用人", "料號", "使用數量", "日期", "時間", "入庫位置"):
        tree_history.heading(col, text=col)
    
    scrollbar = ttk.Scrollbar(search_win, orient="vertical", command=tree_history.yview)
    tree_history.configure(yscrollcommand=scrollbar.set)
    
    tree_history.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def do_search():
        try:
            wb = load_workbook(excel_file)
            ws = wb["使用者入庫"]
            for i in tree_history.get_children():
                tree_history.delete(i)
                
            part_search = search_part.get().strip()
            user_search = search_user.get().strip()
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                user, part, remain, date, time, used, location = row
                # 如果沒有輸入搜尋條件，顯示所有記錄
                if (not part_search and not user_search) or \
                   (part_search and part_search in part) or \
                   (user_search and user_search in user):
                    tree_history.insert("", tk.END, values=(user, part, used, date, time, location))
        except Exception as e:
            messagebox.showerror("錯誤", f"查詢錯誤:\n{e}")

    tk.Button(frame, text="查詢", command=do_search, width=10).grid(row=0, column=4, padx=10)
    tk.Button(frame, text="關閉", command=search_win.destroy, width=10).grid(row=0, column=5, padx=5)

    # 初始顯示所有記錄
    do_search()

# 查詢借出紀錄
def search_lend_history():
    lend_win = tk.Toplevel(root)
    lend_win.title("借出記錄查詢")
    lend_win.geometry("1200x550")

    frame = tk.Frame(lend_win)
    frame.pack(pady=10)

    tk.Label(frame, text="料號:").grid(row=0, column=0, padx=5)
    search_part = tk.Entry(frame, width=20)
    search_part.grid(row=0, column=1, padx=5)

    tk.Label(frame, text="借出人:").grid(row=0, column=2, padx=5)
    search_lender = tk.Entry(frame, width=20)
    search_lender.grid(row=0, column=3, padx=5)

    tree_lend = ttk.Treeview(lend_win, columns=("料號", "庫存數量", "借出狀態", "借出人", "庫存位置"), show="headings")
    tree_lend.column("料號", width=250)
    tree_lend.column("庫存數量", width=100)
    tree_lend.column("借出狀態", width=100)
    tree_lend.column("借出人", width=150)
    tree_lend.column("庫存位置", width=250)
    
    for col in ("料號", "庫存數量", "借出狀態", "借出人", "庫存位置"):
        tree_lend.heading(col, text=col)
    
    scrollbar = ttk.Scrollbar(lend_win, orient="vertical", command=tree_lend.yview)
    tree_lend.configure(yscrollcommand=scrollbar.set)
    
    tree_lend.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def do_search():
        try:
            wb = load_workbook(excel_file)
            ws_inv = wb["目前庫存"]
            for i in tree_lend.get_children():
                tree_lend.delete(i)
                
            part_search = search_part.get().strip()
            lender_search = search_lender.get().strip()
            
            for row in ws_inv.iter_rows(min_row=2, values_only=True):
                part_no, qty, status, lender, location = row
                if status == "借出中":
                    if ((not part_search or part_search in part_no) and
                        (not lender_search or lender_search in lender)):
                        tree_lend.insert("", tk.END, values=(part_no, qty, status, lender, location))
        except Exception as e:
            messagebox.showerror("錯誤", f"讀取借出紀錄錯誤:\n{e}")

    tk.Button(frame, text="查詢", command=do_search, width=10).grid(row=0, column=4, padx=10)
    tk.Button(frame, text="關閉", command=lend_win.destroy, width=10).grid(row=0, column=5, padx=5)

    # 初始顯示所有借出記錄
    do_search()

# 顯示目前庫存獨立視窗
def show_inventory_window():
    inv_win = tk.Toplevel(root)
    inv_win.title("目前庫存清單")
    inv_win.geometry("1000x800")

    frame = tk.Frame(inv_win)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    tree_inv = ttk.Treeview(frame, columns=("料號", "庫存數量", "借出狀態", "借出人", "庫存位置"), show="headings")
    tree_inv.column("料號", width=250)
    tree_inv.column("庫存數量", width=100)
    tree_inv.column("借出狀態", width=100)
    tree_inv.column("借出人", width=150)
    tree_inv.column("庫存位置", width=250)
    
    for col in ("料號", "庫存數量", "借出狀態", "借出人", "庫存位置"):
        tree_inv.heading(col, text=col)
    
    scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree_inv.yview)
    tree_inv.configure(yscrollcommand=scrollbar.set)
    
    tree_inv.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    try:
        wb = load_workbook(excel_file)
        ws_inv = wb["目前庫存"]
        for row in ws_inv.iter_rows(min_row=2, values_only=True):
            status = row[2] if row[2] else ""
            lender = row[3] if row[3] else ""
            location = row[4] if row[4] else ""
            tree_inv.insert("", tk.END, values=(row[0], row[1], status, lender, location))
    except Exception as e:
        messagebox.showerror("錯誤", f"讀取庫存錯誤:\n{e}")
    
    # 關閉視窗按鈕
    btn_close = tk.Button(inv_win, text="關閉", command=inv_win.destroy, width=15)
    btn_close.pack(pady=10)

# 自動刷新 GUI，保持與 Excel 同步
def auto_refresh():
    try:
        refresh_inventory()
    except Exception:
        pass
    root.after(6000, auto_refresh)  # 每6秒刷新一次

# ====== GUI 介面 ======

root = tk.Tk()
root.title("輕量庫存管理系統")
root.geometry("1200x720")

frame_top = tk.Frame(root)
frame_top.pack(pady=10)

tk.Button(frame_top, text="匯入 Excel", command=import_excel, width=15).pack(side=tk.LEFT, padx=10)
tk.Button(frame_top, text="匯出 Excel", command=export_excel, width=15).pack(side=tk.LEFT, padx=10)

frame_in = tk.LabelFrame(root, text="新品入庫 (組別功能)")
frame_in.pack(fill="x", padx=10, pady=5)

tk.Label(frame_in, text="料號:").grid(row=0, column=0, padx=5, pady=5)
entry_in_part = tk.Entry(frame_in, width=20)
entry_in_part.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame_in, text="每組數量:").grid(row=0, column=2, padx=5, pady=5)
entry_in_unit_qty = tk.Entry(frame_in, width=10)
entry_in_unit_qty.grid(row=0, column=3, padx=5, pady=5)

tk.Label(frame_in, text="組數:").grid(row=0, column=4, padx=5, pady=5)
entry_in_groups = tk.Entry(frame_in, width=10)
entry_in_groups.grid(row=0, column=5, padx=5, pady=5)

tk.Label(frame_in, text="入庫位置(備註):").grid(row=0, column=6, padx=5, pady=5)
entry_in_location = tk.Entry(frame_in, width=20)
entry_in_location.grid(row=0, column=7, padx=5, pady=5)

tk.Button(frame_in, text="儲存入庫", command=save_instock, width=15).grid(row=0, column=8, padx=10, pady=5)

frame_lend = tk.LabelFrame(root, text="借出")
frame_lend.pack(fill="x", padx=10, pady=5)
tk.Label(frame_lend, text="借出人:").grid(row=0, column=0, padx=5, pady=5)
entry_lend_user = tk.Entry(frame_lend, width=20)
entry_lend_user.grid(row=0, column=1, padx=5, pady=5)
tk.Label(frame_lend, text="料號(含流水號):").grid(row=0, column=2, padx=5, pady=5)
entry_lend_part = tk.Entry(frame_lend, width=20)
entry_lend_part.grid(row=0, column=3, padx=5, pady=5)
tk.Button(frame_lend, text="借出確認", command=lend_part, width=15).grid(row=0, column=4, padx=10, pady=5)

frame_out = tk.LabelFrame(root, text="使用人入庫")
frame_out.pack(fill="x", padx=10, pady=5)
tk.Label(frame_out, text="使用人:").grid(row=0, column=0, padx=5, pady=5)
entry_user = tk.Entry(frame_out, width=20)
entry_user.grid(row=0, column=1, padx=5, pady=5)
tk.Label(frame_out, text="料號(含流水號):").grid(row=0, column=2, padx=5, pady=5)
entry_part = tk.Entry(frame_out, width=20)
entry_part.grid(row=0, column=3, padx=5, pady=5)
tk.Label(frame_out, text="剩餘數量:").grid(row=0, column=4, padx=5, pady=5)
entry_qty = tk.Entry(frame_out, width=10)
entry_qty.grid(row=0, column=5, padx=5, pady=5)
tk.Label(frame_out, text="入庫位置(備註):").grid(row=0, column=6, padx=5, pady=5)
entry_location = tk.Entry(frame_out, width=20)
entry_location.grid(row=0, column=7, padx=5, pady=5)

tk.Button(frame_out, text="儲存入庫", command=save_outstock, width=15).grid(row=0, column=8, padx=10, pady=5)

tk.Label(root, text="目前庫存總覽", font=("Arial", 12)).pack(pady=5)

frame_inventory = tk.Frame(root)
frame_inventory.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

tree = ttk.Treeview(frame_inventory, columns=("料號", "庫存數量", "借出狀態", "借出人", "庫存位置"), show="headings", height=15)
tree.column("料號", width=250)
tree.column("庫存數量", width=100)
tree.column("借出狀態", width=100)
tree.column("借出人", width=150)
tree.column("庫存位置", width=250)

for col in ("料號", "庫存數量", "借出狀態", "借出人", "庫存位置"):
    tree.heading(col, text=col)
tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

scrollbar = ttk.Scrollbar(frame_inventory, orient=tk.VERTICAL, command=tree.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

tree.configure(yscrollcommand=scrollbar.set)

frame_bottom = tk.Frame(root)
frame_bottom.pack(pady=10)

tk.Button(frame_bottom, text="查詢使用紀錄", command=search_history, width=15).pack(side=tk.LEFT, padx=10)
tk.Button(frame_bottom, text="查詢借出紀錄", command=search_lend_history, width=15).pack(side=tk.LEFT, padx=10)
tk.Button(frame_bottom, text="目前庫存清單", command=show_inventory_window, width=15).pack(side=tk.LEFT, padx=10)

# ====== 程式啟動初始設定 ======

excel_file = load_last_path()
if not excel_file or not os.path.exists(excel_file):
    excel_file = os.path.join(os.getcwd(), "庫存紀錄.xlsx")
    if not os.path.exists(excel_file):
        init_excel(excel_file)
    save_last_path(excel_file)

refresh_inventory()
auto_refresh()

root.mainloop()